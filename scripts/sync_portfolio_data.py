import json
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "portfolio_template.xlsx"
OUTPUT_PATH = ROOT / "assets" / "portfolio-data.json"


def load_workbook_data(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws_guide, ws_basic, ws_high, ws_skills, ws_proj, ws_exp, ws_edu, ws_cert, ws_links, ws_meta = wb.worksheets

    basic = {}
    for row in ws_basic.iter_rows(min_row=2, values_only=True):
        section, key, value, _ = row
        if not section or not key:
            continue
        basic.setdefault(str(section), {})[str(key)] = value

    highlights = []
    for row in ws_high.iter_rows(min_row=2, values_only=True):
        group, order, content = row
        if group and content:
            highlights.append(
                {
                    "group": str(group),
                    "order": int(order) if isinstance(order, (int, float)) else order,
                    "content": str(content),
                }
            )
    highlights.sort(key=lambda x: x.get("order", 0))

    skills_map = defaultdict(list)
    for row in ws_skills.iter_rows(min_row=2, values_only=True):
        category, tech = row
        if category and tech:
            skills_map[str(category)].append(str(tech))
    skills = [{"category": k, "items": v} for k, v in skills_map.items()]

    projects = []
    for row in ws_proj.iter_rows(min_row=2, values_only=True):
        pid, title, period, summary, stack, h1, h2, h3, repo, demo, doc, featured = row
        if not title:
            continue
        projects.append(
            {
                "id": str(pid) if pid else "",
                "title": str(title),
                "period": str(period) if period else "",
                "description": str(summary) if summary else "",
                "tech": [s.strip() for s in str(stack).split(",")] if stack else [],
                "highlights": [str(x) for x in (h1, h2, h3) if x],
                "repoUrl": str(repo) if repo else "",
                "demoUrl": str(demo) if demo else "",
                "docUrl": str(doc) if doc else "",
                "featured": str(featured).strip().upper() == "Y" if featured is not None else False,
            }
        )

    experience = []
    for row in ws_exp.iter_rows(min_row=2, values_only=True):
        company, role, period, desc, a1, a2, a3 = row
        if company:
            experience.append(
                {
                    "company": str(company),
                    "role": str(role) if role else "",
                    "period": str(period) if period else "",
                    "description": str(desc) if desc else "",
                    "bullets": [str(x) for x in (a1, a2, a3) if x],
                }
            )

    education = []
    for row in ws_edu.iter_rows(min_row=2, values_only=True):
        school, major, period, details = row
        if school:
            education.append(
                {
                    "school": str(school),
                    "major": str(major) if major else "",
                    "period": str(period) if period else "",
                    "details": [x.strip() for x in str(details).split(",")] if details else [],
                }
            )

    certifications = []
    for row in ws_cert.iter_rows(min_row=2, values_only=True):
        name, issuer, date = row
        if name:
            certifications.append(
                {
                    "name": str(name),
                    "issuer": str(issuer) if issuer else "",
                    "date": str(date) if date else "",
                }
            )

    links = {}
    for row in ws_links.iter_rows(min_row=2, values_only=True):
        key, value = row
        if key and value:
            links[str(key)] = str(value)

    meta = {}
    for row in ws_meta.iter_rows(min_row=2, values_only=True):
        key, value = row
        if key is not None:
            meta[str(key)] = value

    site = basic.get("site", {})
    hero = basic.get("hero", {})
    about = basic.get("about", {})

    return {
        "site": {
            "title": str(site.get("title") or "Developer Portfolio"),
            "description": str(site.get("description") or ""),
            "lang": str(site.get("lang") or "ko"),
            "baseUrl": str(site.get("baseUrl") or ""),
            "repoName": str(site.get("repoName") or ""),
            "themeColor": str(site.get("themeColor") or "#1459ba"),
        },
        "profile": {
            "name": str(hero.get("name") or ""),
            "role": str(hero.get("role") or ""),
            "tagline": str(hero.get("tagline") or ""),
            "location": str(hero.get("location") or ""),
            "email": str(hero.get("email") or links.get("email") or ""),
            "resumeUrl": str(hero.get("resumeUrl") or ""),
            "avatarUrl": str(hero.get("avatarUrl") or ""),
            "bio": str(about.get("summary") or ""),
        },
        "highlights": [h["content"] for h in highlights if h.get("group") == "about"],
        "skills": skills,
        "projects": projects,
        "experience": experience,
        "education": education,
        "certifications": certifications,
        "contacts": {
            "email": str(links.get("email") or hero.get("email") or ""),
            "github": str(links.get("github") or ""),
            "blog": str(links.get("blog") or ""),
            "linkedin": str(links.get("linkedin") or ""),
        },
        "meta": {
            "version": meta.get("version", 1),
            "updatedAt": str(meta.get("updatedAt") or ""),
        },
    }


def main() -> None:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Input file not found: {XLSX_PATH}")

    data = load_workbook_data(XLSX_PATH)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
